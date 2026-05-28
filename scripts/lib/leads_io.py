"""Shared lead loading from CSV/XLSX and email enrichment merge for GTM scripts."""

from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent.parent
INPUT_DIR = ROOT / "scripts" / "input"
OUTPUT_DIR = ROOT / "scripts" / "output"
LEADS_CSV = INPUT_DIR / "leads.csv"

SOURCE_ADOBE_SUMMIT = "adobe_summit"
SOURCE_LINKEDIN_VARUN = "linkedin_varun"

LINKEDIN_SLUG_RE = re.compile(r"linkedin\.com/in/([^/?#]+)", re.I)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(handle)]


def normalize_company_key(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


def split_name(full_name: str) -> tuple[str, str]:
    parts = full_name.strip().split(None, 1)
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def normalize_id(raw: Any) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    if "." in text:
        try:
            return str(int(float(text)))
        except ValueError:
            pass
    return text


def is_valid_email_status(status: str) -> bool:
    normalized = status.strip().lower()
    return normalized in {"", "ok", "valid"}


def linkedin_slug_from_url(url: str) -> str:
    url = (url or "").strip()
    match = LINKEDIN_SLUG_RE.search(url)
    if match:
        slug = re.sub(r"[^a-zA-Z0-9_-]", "", match.group(1))
        if slug:
            return slug.lower()
    digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:12]
    return digest


def linkedin_id_from_url(url: str) -> str:
    return f"li_{linkedin_slug_from_url(url)}"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass
class EmailEnrichment:
    email: str
    status: str


def load_email_enrichment(output_dir: Path | None = None) -> dict[str, EmailEnrichment]:
    """Merge verified emails from enrichment output files, keyed by lead id."""
    enrichment: dict[str, EmailEnrichment] = {}
    base = output_dir or OUTPUT_DIR
    if not base.exists():
        return enrichment

    patterns = ("leads_enriched_*.csv", "leads_with_emails_*.csv")
    paths: list[Path] = []
    for pattern in patterns:
        paths.extend(sorted(base.glob(pattern)))

    for path in paths:
        for row in read_csv(path):
            lead_id = normalize_id(row.get("id", ""))
            if not lead_id:
                continue
            status = row.get("email_status") or row.get("amf_status") or ""
            if not is_valid_email_status(status):
                continue
            email = (row.get("email") or row.get("valid_email") or row.get("valid_email_only") or "").strip()
            if email and "@" in email:
                enrichment[lead_id] = EmailEnrichment(email=email, status=status or "valid")
    return enrichment


def _email_from_row(row: dict[str, str]) -> EmailEnrichment | None:
    status = row.get("email_status") or row.get("amf_status") or ""
    if not is_valid_email_status(status):
        return None
    email = (row.get("email") or row.get("valid_email") or row.get("valid_email_only") or "").strip()
    if email and "@" in email:
        return EmailEnrichment(email=email, status=status or "valid")
    return None


def _adobe_row_to_lead(row: dict[str, str], email_info: EmailEnrichment | None) -> dict[str, Any]:
    lead_id = normalize_id(row.get("id", ""))
    first, last = split_name(row.get("name", ""))
    company = row.get("company", "")
    lead_tier = row.get("tier", "")
    score_raw = row.get("score", "")
    score = int(normalize_id(score_raw)) if normalize_id(score_raw).isdigit() else None
    tier_int = int(normalize_id(lead_tier)) if normalize_id(lead_tier).isdigit() else None
    dm_raw = row.get("is_decision_maker", "")
    is_dm = dm_raw in ("1", "1.0", "true", "True")

    return {
        "id": lead_id,
        "full_name": row.get("name", ""),
        "first_name": first,
        "last_name": last,
        "company": company,
        "company_key": normalize_company_key(company) if company else "",
        "title": row.get("title", ""),
        "tier": tier_int,
        "tier_reason": row.get("tier_reason", ""),
        "score": score,
        "is_decision_maker": is_dm,
        "dm_reason": row.get("dm_reason", ""),
        "email": email_info.email if email_info else None,
        "email_status": email_info.status if email_info else None,
        "lead_source": SOURCE_ADOBE_SUMMIT,
        "linkedin_url": None,
    }


def load_adobe_summit_csv_rows(rows: list[dict[str, str]], enrichment: dict[str, EmailEnrichment]) -> list[dict[str, Any]]:
    leads: list[dict[str, Any]] = []
    for row in rows:
        lead_id = normalize_id(row.get("id", ""))
        if not lead_id:
            continue
        email_info = enrichment.get(lead_id) or _email_from_row(row)
        leads.append(_adobe_row_to_lead(row, email_info))
    return leads


def _load_xlsx_sheet_rows(path: Path, sheet_index: int = 0) -> list[dict[str, str]]:
    """Load one sheet as list of dicts; falls back to raw XML if openpyxl fails."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        if wb.sheetnames:
            ws = wb[wb.sheetnames[sheet_index]]
        elif wb.worksheets:
            ws = wb.worksheets[sheet_index]
        else:
            wb.close()
            raise ValueError("openpyxl found no worksheets")
        rows = _xlsx_to_dicts(ws)
        wb.close()
        return rows
    except Exception:
        return _load_xlsx_sheet_rows_xml(path, sheet_index)


def _load_xlsx_sheet_rows_xml(path: Path, sheet_index: int = 0) -> list[dict[str, str]]:
    import xml.etree.ElementTree as ET
    import zipfile

    NS = {"m": "http://purl.oclc.org/ooxml/spreadsheetml/main"}
    NS_LEGACY = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    with zipfile.ZipFile(path) as z:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for ns in (NS, NS_LEGACY):
                for si in root.findall("m:si", ns):
                    texts = [t.text or "" for t in si.findall(".//m:t", ns)]
                    shared.append("".join(texts))
                if shared:
                    break

        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels}

        sheet_el = None
        for ns in (NS, NS_LEGACY):
            sheets = wb.findall(".//m:sheets/m:sheet", ns)
            if sheets:
                sheet_el = sheets[sheet_index]
                break
        if sheet_el is None:
            sheet_path = "xl/worksheets/sheet1.xml"
        else:
            rid = sheet_el.attrib.get("{http://purl.oclc.org/ooxml/officeDocument/2006/relationships}id") or sheet_el.attrib.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            target = rel_map.get(rid, "worksheets/sheet1.xml")
            sheet_path = target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"

        root = ET.fromstring(z.read(sheet_path))

        def cell_val(c: ET.Element, ns: dict[str, str]) -> str:
            t = c.attrib.get("t")
            if t == "inlineStr":
                is_el = c.find("m:is", ns)
                if is_el is not None:
                    texts = [t.text or "" for t in is_el.findall(".//m:t", ns)]
                    return "".join(texts)
            v = c.find("m:v", ns)
            if v is None or v.text is None:
                return ""
            if t == "s":
                return shared[int(v.text)]
            return v.text

        rows_dict: dict[int, dict[str, str]] = {}
        for ns in (NS, NS_LEGACY):
            sheet_rows = root.findall(".//m:sheetData/m:row", ns)
            if not sheet_rows:
                continue
            for row in sheet_rows:
                rnum = int(row.attrib.get("r", 0))
                cells: dict[str, str] = {}
                for c in row.findall("m:c", ns):
                    ref = c.attrib.get("r", "")
                    m = re.match(r"([A-Z]+)(\d+)", ref)
                    if m:
                        cells[m.group(1)] = cell_val(c, ns)
                rows_dict[rnum] = cells
            break

        if not rows_dict:
            return []

        header_row_num = min(rows_dict)
        header = rows_dict[header_row_num]
        cols = sorted(header.keys(), key=lambda x: (len(x), x))
        headers = [header[c] for c in cols]

        rows: list[dict[str, str]] = []
        for rnum in sorted(rows_dict):
            if rnum <= header_row_num:
                continue
            row = rows_dict[rnum]
            if not any(str(v).strip() for v in row.values()):
                continue
            rows.append({headers[i]: _cell_str(row.get(cols[i], "")) for i in range(len(cols))})
        return rows


def load_adobe_summit_xlsx(path: Path) -> list[dict[str, Any]]:
    main_rows = _load_xlsx_sheet_rows(path, 0)
    if not main_rows:
        raise ValueError(f"No rows in first sheet of {path}")

    enrichment: dict[str, EmailEnrichment] = {}
    try:
        enriched_rows = _load_xlsx_sheet_rows(path, 1)
        for row in enriched_rows:
            lead_id = normalize_id(row.get("id", ""))
            if not lead_id:
                continue
            info = _email_from_row(row)
            if info:
                enrichment[lead_id] = info
    except (IndexError, ValueError):
        pass

    file_enrichment = load_email_enrichment()
    enrichment = {**file_enrichment, **enrichment}
    return load_adobe_summit_csv_rows(main_rows, enrichment)


def _xlsx_to_dicts(ws: Any) -> list[dict[str, str]]:
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [_cell_str(h) for h in header_row]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        if not values or not any(v is not None and str(v).strip() for v in values):
            continue
        row = {_cell_str(h): _cell_str(v) for h, v in zip(headers, values) if h}
        rows.append(row)
    return rows


def load_linkedin_connections_xlsx(path: Path) -> list[dict[str, Any]]:
    rows = _load_xlsx_sheet_rows(path, 0)
    if not rows:
        rows = _load_xlsx_sheet_rows_xml(path, 0)
    if not rows:
        raise ValueError(f"No rows in {path}")

    # Normalize header keys (LinkedIn export uses spaced names)
    key_map = {
        "first name": "first_name",
        "last name": "last_name",
        "url": "url",
        "email address": "email",
        "company": "company",
        "position": "title",
    }

    leads: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for raw in rows:
        normalized: dict[str, str] = {}
        for key, value in raw.items():
            normalized[key_map.get(key.lower(), key.lower())] = value

        first = normalized.get("first_name", "")
        last = normalized.get("last_name", "")
        url = normalized.get("url", "")
        company = normalized.get("company", "")
        email = normalized.get("email", "").strip() or None
        title = normalized.get("title", "")

        if not first and not last and not url:
            continue

        lead_id = linkedin_id_from_url(url) if url else f"li_{hashlib.sha256(f'{first}|{last}|{company}'.encode()).hexdigest()[:12]}"
        if lead_id in seen_ids:
            suffix = hashlib.sha256(f"{lead_id}|{first}|{last}".encode()).hexdigest()[:6]
            lead_id = f"{lead_id}_{suffix}"
        seen_ids.add(lead_id)

        full_name = " ".join(part for part in (first, last) if part)

        leads.append(
            {
                "id": lead_id,
                "full_name": full_name,
                "first_name": first,
                "last_name": last,
                "company": company,
                "company_key": normalize_company_key(company) if company else "",
                "title": title,
                "tier": None,
                "tier_reason": None,
                "score": None,
                "is_decision_maker": False,
                "dm_reason": None,
                "email": email if email and "@" in email else None,
                "email_status": "provided" if email and "@" in email else None,
                "lead_source": SOURCE_LINKEDIN_VARUN,
                "linkedin_url": url or None,
            }
        )

    return leads


def load_leads(
    input_path: Path | None = None,
    output_dir: Path | None = None,
    limit: int | None = None,
    tier: int | None = None,
    source: str | None = None,
) -> list[dict[str, Any]]:
    path = (input_path or LEADS_CSV).expanduser().resolve()
    if not path.exists():
        raise ValueError(f"Input file not found: {path}")

    suffix = path.suffix.lower()
    resolved_source = source or SOURCE_ADOBE_SUMMIT

    if suffix in (".xlsx", ".xlsm"):
        if resolved_source == SOURCE_LINKEDIN_VARUN:
            leads = load_linkedin_connections_xlsx(path)
        else:
            leads = load_adobe_summit_xlsx(path)
    else:
        rows = read_csv(path)
        if not rows:
            raise ValueError(f"No rows in {path}")
        enrichment = load_email_enrichment(output_dir)
        leads = load_adobe_summit_csv_rows(rows, enrichment)

    if tier is not None:
        leads = [lead for lead in leads if lead.get("tier") == tier]

    if limit is not None:
        leads = leads[:limit]

    return leads
